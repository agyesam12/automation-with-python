import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import sys

def list_file_columns(file_path):
    """
    List all columns available in an Excel file.
    
    Args:
        file_path: Path to Excel file
    """
    try:
        df = pd.read_excel(file_path)
        print(f"\nColumns in '{file_path}':")
        for idx, col in enumerate(df.columns, 1):
            print(f"  {idx}. {col}")
        return df.columns.tolist()
    except Exception as e:
        print(f"✗ Error reading file: {e}")
        return []

def compare_excel_files(file1_path, file2_path, output_path, match_columns=None):
    """
    Compare two Excel files using multiple columns as composite key.
    
    Args:
        file1_path: Path to first Excel file
        file2_path: Path to second Excel file
        output_path: Path to save the comparison report
        match_columns: List of columns to use as composite key for matching (e.g., ['REGION', 'CONSTITUENCY'])
    """
    
    try:
        # Read Excel files
        df1 = pd.read_excel(file1_path)
        df2 = pd.read_excel(file2_path)
        
        print(f"✓ Loaded {file1_path}: {len(df1)} rows")
        print(f"✓ Loaded {file2_path}: {len(df2)} rows")
        
        if match_columns is None:
            match_columns = ['REGION', 'CONSTITUENCY']
        
        # Ensure all match columns exist in both files
        for col in match_columns:
            if col not in df1.columns or col not in df2.columns:
                print(f"✗ Error: '{col}' column not found in one or both files")
                print("\nAvailable columns:")
                list_file_columns(file1_path)
                list_file_columns(file2_path)
                return
        
        # Create composite key from match columns
        df1['_COMPOSITE_KEY'] = df1[match_columns].astype(str).agg('_'.join, axis=1)
        df2['_COMPOSITE_KEY'] = df2[match_columns].astype(str).agg('_'.join, axis=1)
        
        # Check for duplicates
        duplicates1 = df1['_COMPOSITE_KEY'].duplicated().sum()
        duplicates2 = df2['_COMPOSITE_KEY'].duplicated().sum()
        
        if duplicates1 > 0:
            print(f"\n⚠ WARNING: Found {duplicates1} duplicate REGION+CONSTITUENCY combinations in {file1_path}")
        if duplicates2 > 0:
            print(f"⚠ WARNING: Found {duplicates2} duplicate REGION+CONSTITUENCY combinations in {file2_path}")
        
        # Group by composite key and aggregate (keep first occurrence or summarize)
        # For this comparison, we'll keep the first occurrence
        df1_deduped = df1.drop_duplicates(subset='_COMPOSITE_KEY', keep='first')
        df2_deduped = df2.drop_duplicates(subset='_COMPOSITE_KEY', keep='first')
        
        print(f"\n✓ After deduplication: {len(df1_deduped)} unique records in {file1_path}")
        print(f"✓ After deduplication: {len(df2_deduped)} unique records in {file2_path}")
        
        # Set composite key as index
        df1_indexed = df1_deduped.set_index('_COMPOSITE_KEY')
        df2_indexed = df2_deduped.set_index('_COMPOSITE_KEY')
        
        # Initialize results list
        mismatches = []
        
        # Get all unique keys from both files
        all_keys = set(df1_indexed.index) | set(df2_indexed.index)
        
        print(f"✓ Total unique combinations to compare: {len(all_keys)}")
        
        # Compare each record
        for key in sorted(all_keys):
            in_file1 = key in df1_indexed.index
            in_file2 = key in df2_indexed.index
            
            if in_file1 and in_file2:
                # Record exists in both files - check for differences
                row1 = df1_indexed.loc[key]
                row2 = df2_indexed.loc[key]
                
                # Handle case where loc might still return Series (shouldn't happen after dedup, but safe)
                if isinstance(row1, pd.DataFrame):
                    row1 = row1.iloc[0]
                if isinstance(row2, pd.DataFrame):
                    row2 = row2.iloc[0]
                
                # Get all columns from both files (excluding the composite key)
                all_cols = set(df1_deduped.columns) | set(df2_deduped.columns)
                all_cols.discard('_COMPOSITE_KEY')
                
                for col in sorted(all_cols):
                    val1 = row1.get(col) if col in row1.index else None
                    val2 = row2.get(col) if col in row2.index else None
                    
                    # Convert NaN to string for comparison
                    val1_str = str(val1) if pd.notna(val1) else 'MISSING'
                    val2_str = str(val2) if pd.notna(val2) else 'MISSING'
                    
                    if val1_str != val2_str:
                        mismatches.append({
                            'Match_Key': key,
                            'Field': col,
                            'MIGRATE_Value': val1_str,
                            'POLITICAL_Value': val2_str,
                            'Status': 'VALUE_MISMATCH',
                            'File_Source': 'Both'
                        })
            
            elif in_file1 and not in_file2:
                # Record only in file 1
                row1 = df1_indexed.loc[key]
                if isinstance(row1, pd.DataFrame):
                    row1 = row1.iloc[0]
                    
                for col in sorted(df1_deduped.columns):
                    if col != '_COMPOSITE_KEY':
                        val1 = row1.get(col) if col in row1.index else None
                        val1_str = str(val1) if pd.notna(val1) else 'MISSING'
                        mismatches.append({
                            'Match_Key': key,
                            'Field': col,
                            'MIGRATE_Value': val1_str,
                            'POLITICAL_Value': 'N/A - RECORD_NOT_IN_FILE',
                            'Status': 'MISSING_IN_POLITICAL',
                            'File_Source': 'MIGRATE only'
                        })
            
            else:
                # Record only in file 2
                row2 = df2_indexed.loc[key]
                if isinstance(row2, pd.DataFrame):
                    row2 = row2.iloc[0]
                    
                for col in sorted(df2_deduped.columns):
                    if col != '_COMPOSITE_KEY':
                        val2 = row2.get(col) if col in row2.index else None
                        val2_str = str(val2) if pd.notna(val2) else 'MISSING'
                        mismatches.append({
                            'Match_Key': key,
                            'Field': col,
                            'MIGRATE_Value': 'N/A - RECORD_NOT_IN_FILE',
                            'POLITICAL_Value': val2_str,
                            'Status': 'MISSING_IN_MIGRATE',
                            'File_Source': 'POLITICAL only'
                        })
        
        # Create results DataFrame
        results_df = pd.DataFrame(mismatches)
        
        if len(results_df) == 0:
            print("\n✓ No mismatches found! Both files are identical.")
        else:
            print(f"\n✓ Found {len(results_df)} total discrepancies")
            print("\nDiscrepancy Summary by Type:")
            status_counts = results_df['Status'].value_counts()
            for status, count in status_counts.items():
                print(f"  - {status}: {count}")
            
            print("\nFirst 30 discrepancies:")
            print(results_df.head(30).to_string(index=False))
        
        # Save to Excel with formatting
        results_df.to_excel(output_path, index=False, sheet_name='Discrepancies')
        
        # Apply formatting to output file
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Header formatting
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Color coding for status
        value_mismatch_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        missing_political_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        missing_migrate_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            status_cell = row[4]  # Status column (5th column)
            status = status_cell.value
            
            if status:
                if 'VALUE_MISMATCH' in status:
                    status_cell.fill = value_mismatch_fill
                elif 'MISSING_IN_POLITICAL' in status:
                    status_cell.fill = missing_political_fill
                elif 'MISSING_IN_MIGRATE' in status:
                    status_cell.fill = missing_migrate_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        wb.save(output_path)
        print(f"\n✓ Report saved to: {output_path}")
        
    except FileNotFoundError as e:
        print(f"✗ Error: File not found - {e}")
    except Exception as e:
        print(f"✗ Error: {e}")
        import traceback
        traceback.print_exc()

# Example usage
if __name__ == '__main__':
    # Modify these paths to your actual files
    file1 = 'MIGRATE.xlsx'
    file2 = 'political.xlsx'
    output = 'comparison_report.xlsx'
    
    # First, list the columns to discover what's available
    print("=" * 60)
    print("DISCOVERING COLUMNS IN YOUR FILES")
    print("=" * 60)
    list_file_columns(file1)
    list_file_columns(file2)
    
    # Columns to use for matching records between files
    # Using REGION and CONSTITUENCY as composite key
    match_cols = ['REGION', 'CONSTITUENCY']
    
    print("\n" + "=" * 60)
    print("STARTING COMPARISON")
    print("=" * 60)
    print(f"Matching records by: {match_cols}")
    print("=" * 60)
    compare_excel_files(file1, file2, output, match_columns=match_cols)